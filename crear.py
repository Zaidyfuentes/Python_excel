from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

libro = Workbook()
hoja = libro.active
hoja.title = "Nomina Junio 2026"

# titulo columnas
encabezados = [
    "Documento", "Nombre", "Cargo", "Salario", "Horas Extras",
    "Auxilio Transporte", "Valor Hora", "Valor Hora Extra",
    "Total Horas Extras", "Devengado", "Salud", "Pensión", "Neto a Pagar"
]

for col, h in enumerate(encabezados, 1):
    cell = hoja.cell(1, col, h)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="1F4E78")
    cell.alignment = Alignment(horizontal="center")

# usuarios 
usuarios = [
    [1111, "Juan Perez", "Ingeniero", 3000000, 2],
    [1112, "Maria Gomez", "Analista", 2500000, 3],
    [1113, "Carlos Ruiz", "Desarrollador", 2800000, 1],
    [1114, "Ana Torres", "Diseñadora", 2400000, 4],
    [1115, "Luis Diaz", "Soporte", 2200000, 2],
    [1116, "Laura Castro", "QA", 2600000, 3],
    [1117, "Pedro Rojas", "Arquitecto", 4000000, 5],
    [1118, "Sofia Mendoza", "Scrum Master", 3500000, 2],
    [1119, "Miguel Lopez", "DevOps", 3800000, 3],
    [1110, "Paula Vargas", "Product Owner", 4200000, 1]
]

# operaciones
for fila, u in enumerate(usuarios, start=2):
    documento, nombre, cargo, salario, horas = u

    auxilio = 200000 if salario <= 2847000 else 0
    valor_hora = salario / 240
    valor_hora_extra = valor_hora * 1.25
    total_horas_extra = horas * valor_hora_extra

    devengado = salario + auxilio + total_horas_extra
    salud = devengado * 0.04
    pension = devengado * 0.04
    neto = devengado - salud - pension

    valores = [
        documento, nombre, cargo, salario, horas,
        auxilio, valor_hora, valor_hora_extra,
        total_horas_extra, devengado, salud, pension, neto
    ]

    for col, v in enumerate(valores, 1):
        hoja.cell(fila, col, v)

# formato moneda
for row in hoja.iter_rows(min_row=2, min_col=4, max_col=13):
    for cell in row:
        cell.number_format = '"$"#,##0'

# bordes
border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

for row in hoja.iter_rows(min_row=1, max_row=11, min_col=1, max_col=13):
    for cell in row:
        cell.border = border

# ajuste de columnas
for col in hoja.columns:
    max_length = 0
    col_letter = col[0].column_letter
    for cell in col:
        if cell.value:
            max_length = max(max_length, len(str(cell.value)))
    hoja.column_dimensions[col_letter].width = max_length + 2

# hoja resumen
resumen = libro.create_sheet("Resumen Gerencial")

resumen["A1"] = "Indicador"
resumen["B1"] = "Valor"

resumen["A2"] = "Total Empleados"
resumen["A3"] = "Total Nomina"
resumen["A4"] = "Total salud descontada"
resumen["A5"] = "Total pension descontada"
resumen["A6"] = "Mayor Salario"
resumen["A7"] = "promedio Salarial"

resumen["B2"] = "=COUNTA('Nomina Junio 2026'!A2:A11)"

# Total nómina (NETO A PAGAR)
resumen["B3"] = "=SUM('Nomina Junio 2026'!M2:M11)"

# Total salud descontada
resumen["B4"] = "=SUM('Nomina Junio 2026'!K2:K11)"

# Total pensión descontada
resumen["B5"] = "=SUM('Nomina Junio 2026'!L2:L11)"

# Mayor salario
resumen["B6"] = "=MAX('Nomina Junio 2026'!D2:D11)"

# Promedio salarial
resumen["B7"] = "=AVERAGE('Nomina Junio 2026'!D2:D11)"


libro.save("nomina_empresa.xlsx")
print("Archivo generado correctamente")

#activar el .venv : .venv\Scripts\activate
#extension para generar .exe: pip install pyinstaller
#generar .exe: pyinstaller --onefile --noconsole crear.py